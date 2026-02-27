from flask import Flask, render_template, request, redirect, url_for, Response, flash, send_file, render_template
from models import db, Expert, Committee, Membership, Meeting, Participation, NationalMirrorCommittee
from apscheduler.schedulers.background import BackgroundScheduler
import csv
from datetime import date, datetime
from openpyxl import Workbook
from flask_mail import Mail, Message
import io
from emails import (
    announcement_email,
    reminder_email_all,
    reminder_email_individual,
    completion_email,
    request_update_email
)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///committee.db'
app.config['SECRET_KEY'] = 'yoursecretkey'
db.init_app(app)


@app.context_processor
def inject_current_year():
    return {'current_year': date.today().year}

@app.route('/')
def homepage():
    today = date.today()

    nmcs = NationalMirrorCommittee.query.all()
    memberships = Membership.query.all()
    experts = Expert.query.all()
    meetings = Meeting.query.filter(Meeting.date >= today).order_by(Meeting.date.asc()).all()
    participations = Participation.query.all()

    # Totals
    total_nmc = len(nmcs)
    total_sc = sum(len([sc for sc in nmc.subcommittees if sc.parent_id is None]) for nmc in nmcs)
    total_wg = sum(len([wg for wg in nmc.subcommittees if wg.parent_id is not None]) for nmc in nmcs)
    total_memberships = len(memberships)
    total_experts = len(experts)
    unique_expert_ids = {m.expert_id for m in memberships}
    total_unique_experts = len(unique_expert_ids)

    # Build summary for each NMC
    nmc_summary = []
    for nmc in nmcs:
        expert_ids = {m.expert_id for m in memberships if m.committee and m.committee.nmc_id == nmc.id}
        scs = [sc for sc in nmc.subcommittees if sc.parent_id is None]
        wg_count = len([wg for wg in nmc.subcommittees if wg.parent_id is not None])

        sc_summary = []
        for sc in scs:
            sc_expert_ids = {m.expert_id for m in memberships if m.committee_id == sc.id}
            wg_summary = []
            for wg in sc.children:
                wg_expert_ids = {m.expert_id for m in memberships if m.committee_id == wg.id}
                wg_summary.append({
                    "wg": wg,
                    "expert_count": len(wg_expert_ids)
                })
            sc_summary.append({
                "sc": sc,
                "wg_count": len(sc.children),
                "expert_count": len(sc_expert_ids),
                "wgs": wg_summary
            })

        nmc_summary.append({
            "nmc": nmc,
            "sc_count": len(scs),
            "wg_count": wg_count,
            "expert_count": len(expert_ids),
            "scs": sc_summary
        })

    return render_template(
        'index.html',
        nmcs=nmcs,
        memberships=memberships,
        experts=experts,
        meetings=meetings,
        participations=participations,
        nmc_summary=nmc_summary,
        total_nmc=total_nmc,
        total_sc=total_sc,
        total_wg=total_wg,
        total_memberships=total_memberships,
        total_experts=total_experts,
        total_unique_experts=total_unique_experts,
        current_year=today.year
    )

@app.route('/dashboard')
def dashboard():
    nmcs = NationalMirrorCommittee.query.all()
    today = date.today()
    nmc_meetings = {}
    upcoming_count = 0
    past_count = 0

    for nmc in nmcs:
        meetings_set = set()
        for sc in nmc.subcommittees:
            meetings_set.update(sc.meetings)
            for wg in sc.children:
                meetings_set.update(wg.meetings)

        upcoming = [m for m in meetings_set if m.date >= today]
        past = [m for m in meetings_set if m.date < today]

        upcoming_count += len(upcoming)
        past_count += len(past)

        nmc_meetings[nmc.id] = {
            "upcoming_preview": sorted(upcoming, key=lambda m: m.date)[:2],
            "upcoming_all": sorted(upcoming, key=lambda m: m.date),
            "past_preview": sorted(past, key=lambda m: m.date, reverse=True)[:2],
            "past_all": sorted(past, key=lambda m: m.date, reverse=True)
        }

    participations = Participation.query.all()
    return render_template(
        'dashboard.html',
        nmcs=nmcs,
        nmc_meetings=nmc_meetings,
        participations=participations,
        upcoming_count=upcoming_count,
        past_count=past_count
    )
    
@app.route('/directory', methods=['GET', 'POST'])
def directory():
    if request.method == 'POST':
        file = request.files.get('file')
        if file and file.filename.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                nmc_code, nmc_title, sc_code, sc_title, wg_code, wg_title, expert_name, organisation, email, phone = row

                if not email:
                    continue

                # Ensure NMC exists
                nmc = None
                if nmc_code and nmc_title:
                    nmc = NationalMirrorCommittee.query.filter_by(code=nmc_code).first()
                    if not nmc:
                        nmc = NationalMirrorCommittee(code=nmc_code, title=nmc_title)
                        db.session.add(nmc)
                        db.session.flush()

                # Ensure SC exists
                sc = None
                if sc_code and sc_title and nmc:
                    sc = Committee.query.filter_by(code=sc_code).first()
                    if not sc:
                        sc = Committee(code=sc_code, title=sc_title, parent_id=None, nmc_id=nmc.id)
                        db.session.add(sc)
                        db.session.flush()

                # Ensure WG exists
                wg = None
                if wg_code and wg_title and sc:
                    full_wg_code = f"{sc.code}/{wg_code}"
                    wg = Committee.query.filter_by(code=full_wg_code).first()
                    if not wg:
                        wg = Committee(
                            code=full_wg_code,
                            title=wg_title,
                            parent_id=sc.id,
                            nmc_id=sc.nmc_id
                        )
                        db.session.add(wg)
                        db.session.flush()

                # Ensure Expert exists
                expert = Expert.query.filter_by(email=email).first()
                if not expert:
                    expert = Expert(
                        name=expert_name,
                        organisation=organisation,
                        email=email,
                        mobile=phone,
                        is_active=True   # ensure new experts are active
                    )
                    db.session.add(expert)
                    db.session.flush()

                # Link Expert to WG if present, otherwise SC
                target_committee = wg if wg else sc
                if target_committee:
                    membership = Membership.query.filter_by(
                        expert_id=expert.id,
                        committee_id=target_committee.id
                    ).first()
                    if not membership:
                        db.session.add(Membership(expert_id=expert.id, committee_id=target_committee.id))

            db.session.commit()
            flash("Directory updated successfully!", "success")
            return redirect(url_for('directory'))

    nmcs = NationalMirrorCommittee.query.all()
    scs = Committee.query.filter(Committee.parent_id == None).all()
    wgs = Committee.query.filter(Committee.parent_id != None).all()
    experts = Expert.query.filter_by(is_active=True).all()   # only active experts
    memberships = Membership.query.all()

    return render_template(
        'directory.html',
        nmcs=nmcs,
        scs=scs,
        wgs=wgs,
        experts=experts,
        memberships=memberships
    )
    
@app.route('/add_nmc', methods=['GET', 'POST'])
def add_nmc():
    if request.method == 'POST':
        code = request.form['code']
        title = request.form['title']
        nmc = NationalMirrorCommittee(code=code, title=title)
        db.session.add(nmc)
        db.session.commit()
        flash('NMC added successfully!', 'success')
        return redirect(url_for('directory'))
    return render_template('add_nmc.html')

@app.route('/add_sc', methods=['GET', 'POST'])
def add_sc():
    nmcs = NationalMirrorCommittee.query.all()
    if request.method == 'POST':
        code = request.form['code']
        title = request.form['title']
        nmc_id = request.form['nmc_id']
        sc = Committee(code=code, title=title, nmc_id=nmc_id)
        db.session.add(sc)
        db.session.commit()
        flash('SC added successfully!', 'success')
        return redirect(url_for('directory'))
    return render_template('add_sc.html', nmcs=nmcs)

@app.route('/add_wg', methods=['GET', 'POST'])
def add_wg():
    nmcs = NationalMirrorCommittee.query.all()
    scs = Committee.query.filter(Committee.parent_id == None).all()  # only SCs

    if request.method == 'POST':
        code = request.form['code']          # e.g. "WG 9"
        title = request.form['title']
        sc_id = request.form['sc_id']

        # Fetch the SC object
        sc = Committee.query.get(sc_id)
        if not sc:
            flash("Invalid SC selected", "danger")
            return redirect(url_for('directory'))

        # Auto‑prefix SC code to WG code
        merged_code = f"{sc.code}/{code}"

        wg = Committee(
            code=merged_code,
            title=title,
            parent_id=sc.id,     # ✅ link WG to SC
            nmc_id=sc.nmc_id     # ✅ inherit NMC from SC
        )
        db.session.add(wg)
        db.session.commit()

        flash('WG added successfully!', 'success')
        return redirect(url_for('directory'))

    return render_template('add_wg.html', nmcs=nmcs, scs=scs)

@app.route('/add_membership', methods=['GET', 'POST'])
def add_membership():
    nmcs = NationalMirrorCommittee.query.all()
    experts = Expert.query.all()
    scs = Committee.query.filter(Committee.parent_id == None).all()
    wgs = Committee.query.filter(Committee.parent_id != None).all()

    if request.method == 'POST':
        expert_id = request.form['expert_id']
        committee_id = request.form['committee_id']

        membership = Membership(expert_id=expert_id, committee_id=committee_id)
        db.session.add(membership)
        db.session.commit()

        flash('Membership added successfully!', 'success')
        return redirect(url_for('directory'))

    return render_template('add_membership.html', nmcs=nmcs, experts=experts, scs=scs, wgs=wgs)

@app.route('/delete_membership/<int:membership_id>', methods=['POST'])
def delete_membership(membership_id):
    membership = Membership.query.get_or_404(membership_id)
    db.session.delete(membership)
    db.session.commit()
    flash('Membership removed successfully!', 'warning')
    return redirect(url_for('directory'))

# Add Expert
@app.route('/add_expert', methods=['POST'])
def add_expert():
    name = request.form['name']
    email = request.form['email']
    mobile = request.form.get('mobile')
    organisation = request.form.get('organisation')

    new_expert = Expert(
        name=name,
        email=email,
        mobile=mobile,
        organisation=organisation
    )
    db.session.add(new_expert)
    db.session.commit()

    flash('Expert added successfully!', 'success')
    # ✅ Redirect back to Update Directory page
    return redirect(url_for('directory'))

@app.route('/get_scs/<int:nmc_id>')
def get_scs(nmc_id):
    scs = Committee.query.filter_by(nmc_id=nmc_id, parent_id=None).all()
    return jsonify([{'id': sc.id, 'code': sc.code, 'title': sc.title} for sc in scs])

@app.route('/get_wgs/<int:sc_id>')
def get_wgs(sc_id):
    sc = Committee.query.get(sc_id)
    wgs = sc.children if sc else []
    return jsonify([{'id': wg.id, 'code': wg.code, 'title': wg.title} for wg in wgs])

# Edit Expert
@app.route('/edit_expert/<int:expert_id>', methods=['GET', 'POST'])
def edit_expert(expert_id):
    expert = Expert.query.get_or_404(expert_id)
    committees = Committee.query.all()

    if request.method == 'POST':
        expert.name = request.form.get('name')
        expert.email = request.form.get('email')
        expert.mobile = request.form.get('mobile')
        expert.organisation = request.form.get('organisation')

        # Handle staged deletions
        delete_ids = request.form.getlist('delete_memberships')
        for mid in delete_ids:
            membership = Membership.query.get(mid)
            if membership and membership.expert_id == expert.id:
                db.session.delete(membership)

        db.session.commit()
        flash('Expert updated successfully!', 'success')
        return redirect(url_for('directory'))

    # GET request → show edit form
    return render_template('edit_expert.html', expert=expert, committees=committees)

# Delete Expert
@app.route('/delete_expert/<int:expert_id>', methods=['POST'])
def delete_expert(expert_id):
    expert = Expert.query.get_or_404(expert_id)
    expert.is_active = False   # soft delete
    db.session.commit()
    flash('Expert marked inactive successfully!', 'warning')
    return redirect(url_for('directory'))


# Add Meeting

app.config['MAIL_SERVER'] = 'smtp.mgovcloud.in'   # or smtp.zoho.in / smtp.zoho.eu depending on your domain
app.config['MAIL_PORT'] = 587                 # TLS port
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_USERNAME'] = 'sailendra@bis.gov.in'   # full Zoho Gov email
app.config['MAIL_PASSWORD'] = '72dNMNWT9fvw'        # the 16-char app password
app.config['MAIL_DEFAULT_SENDER'] = 'sailendra@bis.gov.in'

mail = Mail(app)

# Add Meeting Route
@app.route('/add_meeting', methods=['GET', 'POST'])
def add_meeting():
    if request.method == 'POST':
        committee_id = request.form['committee_id']
        meeting_no = request.form['meeting_no']
        date_str = request.form['date']
        reg_last_date_str = request.form['registration_last_date']
        agenda = request.form['agenda']

        try:
            meeting_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            reg_last_date = datetime.strptime(reg_last_date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("Invalid date format. Please use YYYY-MM-DD.", "danger")
            return redirect(url_for('add_meeting'))

        # Create meeting
        meeting = Meeting(
            committee_id=committee_id,
            meeting_no=meeting_no,
            date=meeting_date,
            registration_last_date=reg_last_date,
            agenda=agenda
        )
        db.session.add(meeting)
        db.session.commit()

        # Create participations
        committee = Committee.query.get(committee_id)
        memberships = []

        if committee.parent_id is None:
            # SC meeting → only SC experts
            memberships = Membership.query.filter_by(committee_id=committee.id).all()
        else:
            # WG meeting → only WG experts
            memberships = Membership.query.filter_by(committee_id=committee.id).all()

        for m in memberships:
            participation = Participation(meeting_id=meeting.id, expert_id=m.expert_id)
            db.session.add(participation)
        db.session.commit()

        # Send announcement email
        participations = Participation.query.filter_by(meeting_id=meeting.id).all()
        recipient_emails = [p.expert.email for p in participations]
        msg = announcement_email(meeting, recipient_emails)
        mail.send(msg)

        flash('Meeting scheduled, participation prepared, and announcement email sent!', 'success')
        return redirect(url_for('add_meeting'))

    # Prepare upcoming/past meetings grouped by SC/WG
    scs = Committee.query.filter_by(parent_id=None).all()
    nmcs = NationalMirrorCommittee.query.all()
    today = date.today()
    nmc_meetings = {}

    for nmc in nmcs:
        nmc_meetings[nmc.id] = {}

        # SCs
        for sc in nmc.subcommittees:
            upcoming = sorted([m for m in sc.meetings if m.date >= today], key=lambda m: m.date)
            past = sorted([m for m in sc.meetings if m.date < today], key=lambda m: m.date, reverse=True)

            nmc_meetings[nmc.id][sc.code] = {
                "title": sc.title,
                "upcoming_one": upcoming[0] if upcoming else None,
                "past_one": past[0] if past else None,
                "upcoming_all": upcoming,
                "past_all": past
            }

            # WGs under SC
            for wg in sc.children:
                upcoming = sorted([m for m in wg.meetings if m.date >= today], key=lambda m: m.date)
                past = sorted([m for m in wg.meetings if m.date < today], key=lambda m: m.date, reverse=True)

                nmc_meetings[nmc.id][wg.code] = {
                    "title": wg.title,
                    "upcoming_one": upcoming[0] if upcoming else None,
                    "past_one": past[0] if past else None,
                    "upcoming_all": upcoming,
                    "past_all": past
                }

    return render_template(
        'meetings.html',
        scs=scs,
        nmcs=nmcs,
        nmc_meetings=nmc_meetings,
        participations=Participation.query.all()
    )
        
@app.route('/import_meetings', methods=['POST'])
def import_meetings():
    file = request.files.get('excelFile')
    if not file or not file.filename.endswith('.xlsx'):
        flash("No Excel file uploaded", "danger")
        return redirect(url_for('add_meeting'))

    from openpyxl import load_workbook
    wb = load_workbook(file)
    ws = wb.active

    # Expected columns: NMC, SC, WG, Meeting No, Meeting Date, Registration Last Date, Agenda
    for row in ws.iter_rows(min_row=2, values_only=True):
        nmc_code, sc_code, wg_code, meeting_no, meeting_date_val, reg_date_val, agenda = row

        if not sc_code and not wg_code:
            continue

        # --- Meeting Date ---
        if isinstance(meeting_date_val, datetime):
            meeting_date = meeting_date_val.date()
        else:
            try:
                meeting_date = datetime.strptime(str(meeting_date_val).strip(), "%d-%m-%Y").date()
            except Exception:
                flash(f"Invalid meeting date format for SC={sc_code}, WG={wg_code}", "danger")
                continue

        # --- Registration Last Date ---
        reg_last_date = None
        if reg_date_val:
            if isinstance(reg_date_val, datetime):
                reg_last_date = reg_date_val.date()
            else:
                try:
                    reg_last_date = datetime.strptime(str(reg_date_val).strip(), "%d-%m-%Y").date()
                except Exception:
                    flash(f"Invalid registration date format for SC={sc_code}, WG={wg_code}", "warning")

        # --- Committee Lookup ---
        committee = None
        if wg_code and str(wg_code).strip() != "":
            # WG code is just "WG 3" etc., so append to SC code
            full_wg_code = f"{sc_code.strip()}/{wg_code.strip()}"
            committee = Committee.query.filter_by(code=full_wg_code).first()
        else:
            # SC meeting
            committee = Committee.query.filter_by(code=str(sc_code).strip()).first()

        if not committee:
            flash(f"Committee not found for SC={sc_code}, WG={wg_code}", "danger")
            continue

        # --- Create Meeting ---
        meeting = Meeting(
            committee_id=committee.id,
            meeting_no=str(meeting_no),
            date=meeting_date,
            registration_last_date=reg_last_date if reg_last_date else meeting_date,
            agenda=agenda
        )
        db.session.add(meeting)
        db.session.flush()

        # --- Create Participations ---
        # Rule: SC meeting → SC experts only; WG meeting → WG experts only
        memberships = Membership.query.filter_by(committee_id=committee.id).all()

        for m in memberships:
            participation = Participation(meeting_id=meeting.id, expert_id=m.expert_id)
            db.session.add(participation)

    db.session.commit()
    flash("Meetings imported successfully!", "success")
    return redirect(url_for('add_meeting'))

@app.route('/meetings/<path:committee_code>')
def view_all_meetings(committee_code):
    # Find the committee by its code (SC or WG)
    committee = Committee.query.filter_by(code=committee_code).first_or_404()

    today = date.today()

    # Separate upcoming and past meetings
    upcoming_meetings = Meeting.query.filter(
        Meeting.committee_id == committee.id,
        Meeting.date >= today
    ).order_by(Meeting.date.asc()).all()

    past_meetings = Meeting.query.filter(
        Meeting.committee_id == committee.id,
        Meeting.date < today
    ).order_by(Meeting.date.desc()).all()

    # Collect participations for all meetings
    participations = Participation.query.filter(
        Participation.meeting_id.in_([m.id for m in upcoming_meetings + past_meetings])
    ).all()

    return render_template(
        'all_meetings.html',
        committee=committee,
        upcoming_meetings=upcoming_meetings,
        past_meetings=past_meetings,
        participations=participations
    )



# Send Reminder to all participants
@app.route('/send_reminder/<int:meeting_id>', methods=['POST'])
def send_reminder(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    participations = Participation.query.filter_by(meeting_id=meeting.id).all()
    recipient_emails = [p.expert.email for p in participations]

    msg = reminder_email_all(meeting, recipient_emails)
    mail.send(msg)

    for p in participations:
        p.reminder_sent = True
    db.session.commit()

    flash('Reminder sent to all members in one email!', 'info')
    return redirect(url_for('dashboard'))


# Send individual reminder
@app.route('/send_individual_reminder/<int:participation_id>', methods=['POST'])
def send_individual_reminder(participation_id):
    p = Participation.query.get_or_404(participation_id)
    meeting = p.meeting

    msg = reminder_email_individual(meeting, p.expert)
    mail.send(msg)

    p.reminder_sent = True
    db.session.commit()

    flash(f'Reminder sent to {p.expert.name}!', 'info')
    return redirect(url_for('dashboard'))


# Send completion email (after meeting)
@app.route('/send_completion/<int:meeting_id>', methods=['POST'])
def send_completion(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    participations = Participation.query.filter_by(meeting_id=meeting.id).all()
    recipient_emails = [p.expert.email for p in participations]

    msg = completion_email(meeting, recipient_emails)
    mail.send(msg)

    flash('Completion email sent to all experts!', 'success')
    return redirect(url_for('dashboard'))


# Send request update email (to individual expert after completion)
@app.route('/send_request_update/<int:participation_id>', methods=['POST'])
def send_request_update_route(participation_id):
    p = Participation.query.get_or_404(participation_id)
    meeting = p.meeting

    msg = request_update_email(meeting, p.expert)
    mail.send(msg)

    flash(f'Participation status request sent to {p.expert.name}!', 'info')
    return redirect(url_for('dashboard'))

# Add Participation
@app.route('/add_participation/<int:meeting_id>', methods=['POST'])
def add_participation(meeting_id):
    expert_id = request.form['expert_id']
    attendance = 'attendance' in request.form
    report_submitted = 'report_submitted' in request.form
    reminder_sent = 'reminder_sent' in request.form
    participation = Participation(
        meeting_id=meeting_id,
        expert_id=expert_id,
        attendance=attendance,
        report_submitted=report_submitted,
        reminder_sent=reminder_sent
    )
    db.session.add(participation)
    db.session.commit()
    flash('Participation added!', 'success')
    return redirect(url_for('dashboard'))

# Update Participation
@app.route('/update_participation/<int:participation_id>', methods=['POST'])
def update_participation(participation_id):
    participation = Participation.query.get_or_404(participation_id)
    participation.attendance = 'attendance' in request.form
    participation.report_submitted = 'report_submitted' in request.form
    participation.reminder_sent = 'reminder_sent' in request.form
    db.session.commit()
    flash('Participation updated!', 'info')
    # Redirect back to dashboard with past tab active
    return redirect(url_for('dashboard', tab='past'))

# Delete Participation
@app.route('/delete_participation/<int:participation_id>', methods=['POST'])
def delete_participation(participation_id):
    participation = Participation.query.get_or_404(participation_id)
    db.session.delete(participation)
    db.session.commit()
    flash('Participation deleted!', 'danger')
    return redirect(url_for('dashboard'))

# Export for a single committee (NMC)
@app.route('/export_participation/<int:committee_id>')
def export_participation(committee_id):
    committee = NationalMirrorCommittee.query.get_or_404(committee_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Participation Report"
    # Add export date at the very top
    ws.append([f"Export Date: {date.today().strftime('%Y-%m-%d')}"])
    ws.append([])  # blank row for spacing
    # Header row
    ws.append([
        "NMC Code", "NMC Title",
        "SC Code", "SC Title",
        "WG Code", "WG Title",
        "Meeting Date", "Agenda",
        "Expert", "Organisation", "Email", "Phone",
        "Participation", "Report Submitted", "Reminder Sent"
    ])

    for sc in committee.subcommittees:
        for wg in sc.children:
            for meeting in wg.meetings:
                
                if meeting.date < date.today():
                    participations = Participation.query.filter_by(meeting_id=meeting.id).all()
                    for p in participations:
                        ws.append([
                            committee.code,
                            committee.title,
                            sc.code,
                            sc.title,
                            wg.code,
                            wg.title,
                            meeting.date.strftime("%Y-%m-%d"),
                            meeting.agenda or "",
                            p.expert.name,
                            p.expert.organisation or "",
                            p.expert.email or "",
                            p.expert.mobile or "",
                            "Yes" if p.attendance else "No",
                            "Yes" if p.report_submitted else "No",
                            "Yes" if p.reminder_sent else "No"
                        ])

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # ✅ Add export date to filename
    export_date = date.today().strftime("%Y-%m-%d")
    filename = f"{committee.code}_participation_{export_date}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Export for ALL committees
@app.route('/export_all_participation')
def export_all_participation():
    nmcs = NationalMirrorCommittee.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "All Participation"
    # Add export date at the very top
    ws.append([f"Export Date: {date.today().strftime('%Y-%m-%d')}"])
    ws.append([])  # blank row for spacing
    # Header row
    ws.append([
        "NMC Code", "NMC Title",
        "SC Code", "SC Title",
        "WG Code", "WG Title",
        "Meeting Date", "Agenda",
        "Expert", "Organisation", "Email", "Phone",
        "Participation", "Report Submitted", "Reminder Sent"
    ])

    for nmc in nmcs:
        for sc in nmc.subcommittees:
            for wg in sc.children:
                for meeting in wg.meetings:
                    if meeting.date < date.today():
                        participations = Participation.query.filter_by(meeting_id=meeting.id).all()
                        for p in participations:
                            ws.append([
                                nmc.code,
                                nmc.title,
                                sc.code,
                                sc.title,
                                wg.code,
                                wg.title,
                                meeting.date.strftime("%Y-%m-%d"),
                                meeting.agenda or "",
                                p.expert.name,
                                p.expert.organisation or "",
                                p.expert.email or "",
                                p.expert.mobile or "",
                                "Yes" if p.attendance else "No",
                                "Yes" if p.report_submitted else "No",
                                "Yes" if p.reminder_sent else "No"
                            ])

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
     # ✅ Add export date to filename
    export_date = date.today().strftime("%Y-%m-%d")
    filename = f"all_committees_participation_{export_date}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
# export experts with their NMC and SC/WG memberships
@app.route('/export_experts')
def export_experts():
    experts = Expert.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Experts"

    # Header row
    ws.append(["Name", "Email", "Phone", "Organisation", "NMC", "Nominations (SC/WG)"])

    for expert in experts:
        # Group memberships by NMC
        nmc_map = {}
        for m in expert.memberships:
            committee = m.committee
            # Walk up hierarchy until top-level NMC
            while committee.parent_id is not None:
                committee = committee.parent
            nmc_code = committee.nmc.code  # ✅ top-level NMC code
            if nmc_code not in nmc_map:
                nmc_map[nmc_code] = []
            nmc_map[nmc_code].append(f"{m.committee.code} - {m.committee.title}")

        # Write one row per NMC
        if nmc_map:
            for nmc_code, nominations in nmc_map.items():
                ws.append([
                    expert.name,
                    expert.email,
                    expert.mobile or "",
                    expert.organisation or "",
                    nmc_code,
                    "; ".join(nominations)
                ])
        else:
            # Expert with no memberships
            ws.append([
                expert.name,
                expert.email,
                expert.mobile or "",
                expert.organisation or "",
                "None",
                "None"
            ])

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="experts_summary.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
def send_completion_emails():
    today = date.today()
    past_meetings = Meeting.query.filter(
        Meeting.date < today,
        Meeting.completion_sent == False
    ).all()

    for meeting in past_meetings:
        participations = Participation.query.filter_by(meeting_id=meeting.id).all()
        recipient_emails = [p.expert.email for p in participations]

        msg = completion_email(meeting, recipient_emails)
        mail.send(msg)

        meeting.completion_sent = True
        db.session.commit()

        print(f"Completion email sent for meeting {meeting.id}")

def ordinal(value):
    try:
        n = int(value)
    except (ValueError, TypeError):
        return value  # if it can't be converted, just return as-is

    if 11 <= n % 100 <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"

# Register Jinja filter 
app.jinja_env.filters['ordinal'] = ordinal

# --- Scheduler setup ---
scheduler = BackgroundScheduler()
scheduler.add_job(send_completion_emails, 'interval', days=1)  # run once per day
scheduler.start()

if __name__ == "__main__":
    with app.app_context():
        #db.drop_all()
        db.create_all()
    app.run(host="0.0.0.0", port=5001, debug=True)
